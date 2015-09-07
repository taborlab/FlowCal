#!/usr/bin/python
#
# excel_io.py - Module containing functions to read/write excel metadata 
# files.
#
# Authors: Brian Landry (brian.landry@rice.edu)
#          Sebastian M. Castillo-Hair (smc9@rice.edu)
# 
# Date: 7/13/2015

import collections

import xlrd
import openpyxl

def import_rows(workbook_name, worksheet_name):
    ''' Opens an excel doument and imports data from a specific sheet.

    This function returns a list of dictionaries, representing a row each.
    Values in the dictionary will correspond to each cell contents, while
    keys will be taken from the first (header) row.

    The header row is processed until an empty cell is found. Rows are read 
    until a fully empty row is reached. Rows with some fields empty are 
    acceptable.

    Arguments:

    workbook_name   - Name of the excel file to open.
    worksheet_name  - Name of the sheet to read.

    Returns:

    rows_data       - A list of ordered dictionaries, one dictionary per row.
    '''
    rows_data = []
    # Load workbook and sheet
    workbook = xlrd.open_workbook(workbook_name)
    if worksheet_name not in workbook.sheet_names():
        raise IOError('No sheet named {} found'.format(worksheet_name))
    worksheet = workbook.sheet_by_name(worksheet_name)

    for r in range(worksheet.nrows):
        row = worksheet.row(r)

        # Import headers
        if r==0:
            headers = [header.value for header in row]

        # Import row data
        else:
            # Go through row fields
            row_data = collections.OrderedDict()
            for c, cell in enumerate(row):
                value = cell.value
                if value != None and isinstance(value, basestring) \
                    and len(value) > 0 and value[0] == '=':
                    raise ImportError("Excel formula found at {}{}. \
                        Use plain text only.".format(r,c))
                if headers[c] != '':
                    row_data[headers[c]] = value                

            # If row is empty, break
            if sum([v != None for k, v in row_data.iteritems()]) == 0:
                print([v is None for k, v in row_data.iteritems()])
                break
            
            # Save row
            rows_data.append(row_data)

    return rows_data


def export_workbook(workbook_name, worksheet_data):
    ''' Exports data to a workbook.

    worksheet_data is a dictionary of lists of lists. Each dictionary key is 
    used as the worksheet name, and created if necessary. Each list of lists
    corresponds to the rows and columns to save to the worksheet (rows, then 
    columns).

    This function overwrites cells but not worksheets.

    Arguments:
    workbook_name   - Name of the workbook to save data to.
    worksheet_data  - Dictionary of list of lists, representing data to save.
    '''
    # Create Workbook
    workbook = openpyxl.Workbook()
    workbook.remove_sheet(workbook.get_active_sheet())
    
    starting_index = 1 if float(openpyxl.__version__.split('.')[0])>=2 else 0
    for name, data in worksheet_data.iteritems():
        # Add sheet
        workbook.create_sheet().title = name
        sheet = workbook.get_sheet_by_name(name)

        for r, row in enumerate(data):
            for c, value in enumerate(row):
                sheet.cell(row = r + starting_index, column = c + starting_index).value = value

    # Try to save document   
    try:
        workbook.save(workbook_name)
    except IOError:
        raise IOError("Cannot access Excel document. Is the document open?")